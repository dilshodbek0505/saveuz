import base64
import logging
import os
from io import BytesIO
from pathlib import Path

from django.conf import settings

from import_export.formats.base_formats import XLSX
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string, get_column_letter
from PIL import Image as PILImage, UnidentifiedImageError
from tablib import Dataset


class ImageAwareXLSX(XLSX):
    logger = logging.getLogger(__name__)
    title = "xlsx (images)"

    image_column_name = "image"
    export_row_height = 192  # points (~256px)
    export_column_width = 40  # approx characters to fit ~256px
    export_image_max_size = (256, 256)  # pixels
    export_native_extensions = {".png", ".jpg", ".jpeg", ".bmp", ".gif"}
    import_allowed_formats = {"jpeg", "jpg", "png"}

    def export_data(self, dataset, **kwargs):
        content = super().export_data(dataset, **kwargs)

        if not dataset.headers or self.image_column_name not in dataset.headers:
            return content

        image_col_idx = dataset.headers.index(self.image_column_name) + 1
        workbook = load_workbook(filename=BytesIO(content))
        sheet = workbook.active

        for row_number, row in enumerate(dataset.dict, start=2):
            image_value = row.get(self.image_column_name)
            if not image_value:
                continue

            file_path = image_value
            if not os.path.isabs(file_path):
                file_path = os.path.join(settings.MEDIA_ROOT, image_value)

            if not os.path.isfile(file_path):
                continue

            excel_image = self._prepare_excel_image(file_path)
            if excel_image is None:
                continue

            cell = sheet.cell(row=row_number, column=image_col_idx).coordinate
            excel_image.anchor = cell
            sheet.add_image(excel_image)

            sheet.cell(row=row_number, column=image_col_idx, value="")
            sheet.row_dimensions[row_number].height = max(
                sheet.row_dimensions[row_number].height or 0,
                self.export_row_height,
            )
            col_letter = get_column_letter(image_col_idx)
            sheet.column_dimensions[col_letter].width = max(
                sheet.column_dimensions[col_letter].width or 0,
                self.export_column_width,
            )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def create_dataset(self, in_stream, **kwargs):
        self.logger.debug("ImageAwareXLSX.create_dataset: starting, stream type=%s", type(in_stream).__name__)
        if hasattr(in_stream, "read"):
            raw = in_stream.read()
        else:
            raw = in_stream

        if raw is None:
            self.logger.error("ImageAwareXLSX.create_dataset: received no data for import")
            raise TypeError("No data provided for import")

        if isinstance(raw, str):
            raw = raw.encode("utf-8")
        elif isinstance(raw, memoryview):
            raw = raw.tobytes()
        elif isinstance(raw, bytearray):
            raw = bytes(raw)

        if not isinstance(raw, (bytes, bytearray)):
            self.logger.error(
                "ImageAwareXLSX.create_dataset: unexpected data type %s",
                type(raw).__name__,
            )
            raise TypeError("Unsupported data buffer type")

        self.logger.debug("ImageAwareXLSX.create_dataset: buffer length=%s bytes", len(raw))
        dataset = super().create_dataset(BytesIO(raw), **kwargs)
        self.logger.debug(
            "ImageAwareXLSX.create_dataset: dataset created headers=%s count=%s",
            getattr(dataset, "headers", None),
            len(dataset) if hasattr(dataset, "__len__") else "n/a",
        )

        if not isinstance(dataset, Dataset) or not dataset.headers:
            self.logger.debug("ImageAwareXLSX.create_dataset: dataset has no headers, skipping image extraction")
            return dataset

        if self.image_column_name not in dataset.headers:
            self.logger.debug(
                "ImageAwareXLSX.create_dataset: '%s' column not present in headers=%s",
                self.image_column_name,
                dataset.headers,
            )
            return dataset

        image_col_idx = dataset.headers.index(self.image_column_name) + 1
        workbook = load_workbook(filename=BytesIO(raw))
        sheet = workbook.active

        for picture in getattr(sheet, "_images", []):
            anchor = getattr(picture, "anchor", None)
            if anchor is None:
                continue

            row_number = None
            column_number = None

            if hasattr(anchor, "_from"):
                row_number = anchor._from.row + 1
                column_number = anchor._from.col + 1
            elif isinstance(anchor, str):
                from openpyxl.utils.cell import coordinate_from_string

                column_letter, row_number = coordinate_from_string(anchor)
                column_number = column_index_from_string(column_letter)

            if row_number is None or column_number is None:
                continue

            if column_number != image_col_idx:
                continue

            dataset_row_index = row_number - 2  # compensate header
            if not (0 <= dataset_row_index < len(dataset)):
                continue

            image_bytes = picture._data()
            self.logger.debug(
                "ImageAwareXLSX.create_dataset: found image at row=%s column=%s",
                row_number,
                column_number,
            )
            encoded_image = self._encode_import_image(image_bytes)
            if encoded_image is None:
                continue

            dataset[dataset_row_index, image_col_idx - 1] = encoded_image

        self.logger.debug("ImageAwareXLSX.create_dataset: completed image extraction")
        return dataset

    def _prepare_excel_image(self, file_path):
        extension = Path(file_path).suffix.lower()
        try:
            with PILImage.open(file_path) as img:
                img = img.copy()
                if img.mode not in ("RGB", "RGBA"):
                    img = img.convert("RGBA") if img.mode in ("P", "LA") else img.convert("RGB")

                if img.size != self.export_image_max_size:
                    img = img.resize(self.export_image_max_size, PILImage.LANCZOS)

                converted = BytesIO()
                img.save(converted, format="PNG")
                converted.seek(0)

            image = XLImage(converted)
            image.format = "png"
            return image
        except (UnidentifiedImageError, FileNotFoundError, OSError) as exc:
            self.logger.error(
                "ImageAwareXLSX._prepare_excel_image failed for %s: %s",
                file_path,
                exc,
            )
            return None

    def _encode_import_image(self, image_bytes):
        stream = BytesIO(image_bytes)
        try:
            with PILImage.open(stream) as image:
                original_format = (image.format or "PNG").lower()
                self.logger.debug(
                    "ImageAwareXLSX._encode_import_image: original_format=%s size=%s mode=%s",
                    original_format,
                    image.size,
                    image.mode,
                )
                if original_format not in self.import_allowed_formats:
                    image = image.convert("RGBA")
                elif original_format in {"jpeg", "jpg"} and image.mode not in ("RGB", "L"):
                    image = image.convert("RGB")
                elif original_format == "png" and image.mode in ("P", "LA"):
                    image = image.convert("RGBA")

                if image.size != self.export_image_max_size:
                    image = image.resize(self.export_image_max_size, PILImage.LANCZOS)

                if original_format in {"jpeg", "jpg"}:
                    save_format = "JPEG"
                    mime_type = "image/jpeg"
                else:
                    save_format = "PNG"
                    mime_type = "image/png"

                buffer = BytesIO()
                image.save(buffer, format=save_format)
                encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
                self.logger.debug(
                    "ImageAwareXLSX._encode_import_image: encoded length=%s mime=%s",
                    len(encoded),
                    mime_type,
                )
                return f"data:{mime_type};base64,{encoded}"
        except (UnidentifiedImageError, OSError) as exc:
            self.logger.error("ImageAwareXLSX._encode_import_image failed: %s", exc)
            return None
